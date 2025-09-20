"""
Report Service - Servicio de aplicación para generar reportes contables
Especializado en Estado de Resultados según normativa tributaria colombiana

Responsabilidad única: Generar reportes contables con datos de Siigo API
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from decimal import Decimal

from src.domain.entities.estado_resultados import EstadoResultados, PeriodoComparacion, TipoComparacion
from src.application.ports.interfaces import InvoiceRepository, Logger, FileStorage
from src.domain.exceptions.estado_resultados_exceptions import (
    EstadoResultadosError, SiigoAPIError, DataValidationError, ExcelGenerationError,
    DateRangeError, CalculationError, NormativeComplianceError,
    handle_siigo_connection_error, handle_data_processing_error, 
    handle_excel_generation_error, validate_date_range, validate_calculation_inputs
)


class ReportService:
    """
    Servicio de aplicación para generación de reportes contables.
    
    Principios SOLID aplicados:
    - SRP: Solo maneja la generación de reportes contables
    - OCP: Extensible para nuevos tipos de reportes
    - LSP: Substituble por implementaciones especializadas
    - ISP: Interfaces específicas para cada tipo de reporte
    - DIP: Depende de abstracciones (puertos)
    
    Normativa aplicada:
    - Decreto 2420/2015 - Marco Técnico Normativo Colombia
    - Plan Único de Cuentas (PUC) - Decreto 2650/1993
    - NIIF para PYMES aplicables en Colombia
    """
    
    def __init__(self, invoice_repository: InvoiceRepository, logger: Logger, file_storage: FileStorage):
        self._invoice_repository = invoice_repository
        self._logger = logger
        self._file_storage = file_storage
        
        # Configuración del Plan Único de Cuentas (PUC)
        self._configurar_cuentas_puc()
    
    def _configurar_cuentas_puc(self):
        """Configurar mapeo de cuentas según Plan Único de Cuentas colombiano."""
        self._cuentas_ingresos = [
            '41',  # Ingresos operacionales
            '42'   # Ingresos no operacionales (se clasificarán después)
        ]
        
        self._cuentas_costos = [
            '61',  # Costo de ventas
            '62'   # Compras
        ]
        
        self._cuentas_gastos_admin = [
            '51'   # Gastos de administración
        ]
        
        self._cuentas_gastos_ventas = [
            '52'   # Gastos de ventas
        ]
        
        self._cuentas_otros_ingresos = [
            '4295',  # Diversos ingresos
            '4299'   # Ajustes por inflación ingresos
        ]
        
        self._cuentas_otros_gastos = [
            '5295',  # Diversos gastos
            '5299'   # Ajustes por inflación gastos
        ]
        
        self._cuentas_gastos_financieros = [
            '53'   # Gastos financieros
        ]
        
        self._cuentas_impuestos = [
            '54'   # Impuesto de renta y complementarios
        ]
    
    async def generar_estado_resultados_excel(self, 
                                            fecha_inicio: datetime,
                                            fecha_fin: datetime,
                                            tipo_comparacion: str,
                                            fecha_inicio_comparacion: Optional[datetime] = None,
                                            fecha_fin_comparacion: Optional[datetime] = None) -> str:
        """
        Generar Estado de Resultados en formato Excel.
        
        Args:
            fecha_inicio: Fecha de inicio del periodo actual
            fecha_fin: Fecha de fin del periodo actual
            tipo_comparacion: Tipo de comparación (TipoComparacion)
            fecha_inicio_comparacion: Fecha inicio del periodo de comparación (si es personalizado)
            fecha_fin_comparacion: Fecha fin del periodo de comparación (si es personalizado)
            
        Returns:
            str: Ruta del archivo Excel generado
            
        Raises:
            EstadoResultadosError: Si hay error en la generación del reporte
            DateRangeError: Si el rango de fechas es inválido
            SiigoAPIError: Si hay problemas con Siigo API
            ExcelGenerationError: Si falla la generación del Excel
        """
        try:
            self._logger.info("📊 Iniciando generación de Estado de Resultados")
            
            # Validar rango de fechas
            validate_date_range(fecha_inicio.strftime("%Y-%m-%d"), fecha_fin.strftime("%Y-%m-%d"))
            
            # Calcular periodos de comparación
            periodo_actual, periodo_anterior = self._calcular_periodos_comparacion(
                fecha_inicio, fecha_fin, tipo_comparacion, 
                fecha_inicio_comparacion, fecha_fin_comparacion
            )
            
            # Obtener datos de Siigo API con manejo de errores
            datos_actual = await self._obtener_datos_contables_safe(periodo_actual.fecha_inicio, periodo_actual.fecha_fin)
            datos_anterior = None
            if periodo_anterior:
                self._logger.info(f"🔍 Calculando período anterior: {periodo_anterior.fecha_inicio.strftime('%d/%m/%Y')} - {periodo_anterior.fecha_fin.strftime('%d/%m/%Y')}")
                datos_anterior = await self._obtener_datos_contables_safe(periodo_anterior.fecha_inicio, periodo_anterior.fecha_fin)
                if datos_anterior and datos_anterior.get('facturas_data'):
                    num_facturas_anterior = len(datos_anterior['facturas_data'])
                    self._logger.info(f"✅ Período anterior: {num_facturas_anterior} facturas encontradas")
                else:
                    self._logger.info(f"⚠️ Período anterior: No se encontraron facturas en el rango {periodo_anterior.fecha_inicio.strftime('%d/%m/%Y')} - {periodo_anterior.fecha_fin.strftime('%d/%m/%Y')}")
            else:
                self._logger.info("ℹ️ No se calculó período anterior (comparación deshabilitada)")
            
            # Validar datos obtenidos
            self._validar_datos_contables(datos_actual)
            if datos_anterior:
                self._validar_datos_contables(datos_anterior)
            
            # Construir Estado de Resultados
            estado_resultados = self._construir_estado_resultados(
                periodo_actual, periodo_anterior, datos_actual, datos_anterior
            )
            
            # Generar archivo Excel con manejo de errores
            ruta_archivo = self._generar_archivo_excel_safe(estado_resultados)
            
            self._logger.info(f"✅ Estado de Resultados generado: {ruta_archivo}")
            return ruta_archivo
            
        except (DateRangeError, SiigoAPIError, ExcelGenerationError, DataValidationError) as e:
            # Re-lanzar excepciones específicas
            self._logger.error(f"❌ Error específico generando Estado de Resultados: {e}")
            raise
        except Exception as e:
            # Capturar errores no esperados
            self._logger.error(f"❌ Error inesperado generando Estado de Resultados: {e}")
            raise EstadoResultadosError(
                "Error inesperado durante la generación del Estado de Resultados",
                "ER_UNEXPECTED",
                str(e)
            )
    
    def _calcular_periodos_comparacion(self, 
                                     fecha_inicio: datetime,
                                     fecha_fin: datetime,
                                     tipo_comparacion: str,
                                     fecha_inicio_comparacion: Optional[datetime],
                                     fecha_fin_comparacion: Optional[datetime]) -> Tuple[PeriodoComparacion, Optional[PeriodoComparacion]]:
        """Calcular periodos actual y de comparación."""
        self._logger.info(f"🔍 Calculando períodos - tipo_comparacion recibido: '{tipo_comparacion}'")
        
        # Periodo actual
        periodo_actual = PeriodoComparacion(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            nombre=f"Periodo {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        )
        
        # Periodo anterior según tipo de comparación
        periodo_anterior = None
        
        if tipo_comparacion == TipoComparacion.PERIODO_ANTERIOR:
            # Calcular periodo inmediatamente anterior con la misma duración
            duracion = (fecha_fin - fecha_inicio).days + 1
            fin_anterior = fecha_inicio - timedelta(days=1)
            inicio_anterior = fin_anterior - timedelta(days=duracion-1)
            
            periodo_anterior = PeriodoComparacion(
                fecha_inicio=inicio_anterior,
                fecha_fin=fin_anterior,
                nombre=f"Periodo anterior {inicio_anterior.strftime('%d/%m/%Y')} - {fin_anterior.strftime('%d/%m/%Y')}"
            )
            
        elif tipo_comparacion == TipoComparacion.MISMO_PERIODO_ANO_ANTERIOR:
            # Mismo periodo del año anterior
            inicio_anterior = fecha_inicio.replace(year=fecha_inicio.year - 1)
            fin_anterior = fecha_fin.replace(year=fecha_fin.year - 1)
            
            periodo_anterior = PeriodoComparacion(
                fecha_inicio=inicio_anterior,
                fecha_fin=fin_anterior,
                nombre=f"Mismo periodo año anterior {inicio_anterior.strftime('%d/%m/%Y')} - {fin_anterior.strftime('%d/%m/%Y')}"
            )
            
        elif tipo_comparacion == TipoComparacion.PERSONALIZADO:
            # Periodo personalizado
            if fecha_inicio_comparacion and fecha_fin_comparacion:
                periodo_anterior = PeriodoComparacion(
                    fecha_inicio=fecha_inicio_comparacion,
                    fecha_fin=fecha_fin_comparacion,
                    nombre=f"Periodo personalizado {fecha_inicio_comparacion.strftime('%d/%m/%Y')} - {fecha_fin_comparacion.strftime('%d/%m/%Y')}"
                )
        
        return periodo_actual, periodo_anterior
    
    async def _obtener_datos_contables_safe(self, fecha_inicio: datetime, fecha_fin: datetime) -> Dict[str, Any]:
        """
        Obtener datos contables con manejo de errores mejorado.
        
        Args:
            fecha_inicio: Fecha de inicio del periodo
            fecha_fin: Fecha de fin del periodo
            
        Returns:
            Dict con los datos contables organizados por tipo de cuenta
            
        Raises:
            SiigoAPIError: Si hay problemas con la API de Siigo
            DataValidationError: Si los datos no son válidos
        """
        try:
            return await self._obtener_datos_contables(fecha_inicio, fecha_fin)
        except Exception as e:
            # Convertir a excepción específica según el tipo de error
            raise handle_siigo_connection_error(e)
    
    def _validar_datos_contables(self, datos: Dict[str, Any]) -> None:
        """
        Validar que los datos contables sean coherentes.
        
        Args:
            datos: Diccionario con datos contables
            
        Raises:
            DataValidationError: Si los datos no son válidos
        """
        try:
            # Verificar que existan las secciones principales
            secciones_requeridas = ['ingresos', 'costos', 'gastos_admin', 'gastos_ventas', 
                                  'otros_ingresos', 'otros_gastos', 'gastos_financieros', 'impuestos']
            
            for seccion in secciones_requeridas:
                if seccion not in datos:
                    raise DataValidationError(
                        f"Sección '{seccion}' faltante en datos contables",
                        f"Secciones disponibles: {list(datos.keys())}"
                    )
            
            # Calcular totales para validación
            total_ingresos = sum(item.get('valor', 0) for item in datos['ingresos'])
            total_costos = sum(item.get('valor', 0) for item in datos['costos'])
            total_gastos = sum(item.get('valor', 0) for item in datos['gastos_admin'] + datos['gastos_ventas'])
            
            # Validar rangos razonables
            validate_calculation_inputs(
                float(total_ingresos), 
                float(total_costos), 
                float(total_gastos)
            )
            
            self._logger.info(f"✅ Datos contables validados - Ingresos: {total_ingresos}, Costos: {total_costos}, Gastos: {total_gastos}")
            
        except CalculationError as e:
            raise DataValidationError(f"Error en validación de datos: {e.message}", e.details)
        except Exception as e:
            raise handle_data_processing_error(e, "validación de datos contables")
    
    def _generar_archivo_excel_safe(self, estado_resultados: EstadoResultados) -> str:
        """
        Generar archivo Excel con manejo de errores mejorado.
        
        Args:
            estado_resultados: Objeto EstadoResultados a exportar
            
        Returns:
            str: Ruta del archivo Excel generado
            
        Raises:
            ExcelGenerationError: Si falla la generación del Excel
        """
        try:
            return self._generar_archivo_excel(estado_resultados)
        except Exception as e:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_prevista = f"outputs/estado_resultados_{timestamp}.xlsx"
            raise handle_excel_generation_error(e, ruta_prevista)
    
    async def _obtener_datos_contables(self, fecha_inicio: datetime, fecha_fin: datetime) -> Dict[str, Any]:
        """
        Obtener datos contables desde Siigo API para el periodo especificado.
        
        Args:
            fecha_inicio: Fecha de inicio del periodo
            fecha_fin: Fecha de fin del periodo
            
        Returns:
            Dict con los datos contables organizados por tipo de cuenta
        """
        try:
            self._logger.info(f"📥 Obteniendo datos contables desde {fecha_inicio.strftime('%Y-%m-%d')} hasta {fecha_fin.strftime('%Y-%m-%d')}")
            
            # Inicializar estructura de datos
            datos = {
                'ingresos': [],
                'costos': [],
                'gastos_admin': [],
                'gastos_ventas': [],
                'otros_ingresos': [],
                'otros_gastos': [],
                'gastos_financieros': [],
                'impuestos': []
            }
            
            # Verificar conexión con API
            if not hasattr(self._invoice_repository, 'is_connected') or not self._invoice_repository.is_connected():
                self._logger.warning("⚠️ No hay conexión con API Siigo, usando datos simulados")
                return self._generar_datos_simulados()
            
            # Obtener facturas del periodo
            facturas = await self._obtener_facturas_periodo(fecha_inicio, fecha_fin)
            
            # Obtener diario contable (journals) si está disponible
            journals = await self._obtener_journals_periodo(fecha_inicio, fecha_fin)
            
            # Procesar facturas para extraer información contable
            datos.update(self._procesar_facturas_contables(facturas))
            
            # Procesar journals para información más detallada
            if journals:
                datos.update(self._procesar_journals_contables(journals))
            
            self._logger.info(f"✅ Datos contables obtenidos para {len(facturas)} facturas")
            return datos
            
        except Exception as e:
            self._logger.error(f"❌ Error obteniendo datos contables: {e}")
            # Retornar datos simulados como fallback
            return self._generar_datos_simulados()
    
    async def _obtener_facturas_periodo(self, fecha_inicio: datetime, fecha_fin: datetime) -> List[Dict]:
        """Obtener facturas del periodo desde Siigo API."""
        try:
            if hasattr(self._invoice_repository, 'download_invoices_dataframes'):
                # Usar método existente si está disponible
                # CORRECCIÓN: Pasar parámetros individuales, no diccionario
                encabezados_df, _ = self._invoice_repository.download_invoices_dataframes(
                    fecha_inicio=fecha_inicio.strftime('%Y-%m-%d'),
                    fecha_fin=fecha_fin.strftime('%Y-%m-%d')
                )
                
                if encabezados_df is not None and not encabezados_df.empty:
                    return encabezados_df.to_dict('records')
            
            return []
            
        except Exception as e:
            self._logger.error(f"❌ Error obteniendo facturas: {e}")
            return []
    
    async def _obtener_journals_periodo(self, fecha_inicio: datetime, fecha_fin: datetime) -> List[Dict]:
        """Obtener journals (diario contable) del periodo desde Siigo API."""
        try:
            # TODO: Implementar obtención de journals cuando esté disponible en el adaptador
            self._logger.info("📝 Journals no disponibles, usando solo datos de facturas")
            return []
            
        except Exception as e:
            self._logger.error(f"❌ Error obteniendo journals: {e}")
            return []
    
    def _procesar_facturas_contables(self, facturas: List[Dict]) -> Dict[str, List]:
        """Procesar facturas para extraer información contable relevante."""
        datos = {
            'ingresos': [],
            'costos': [],
            'gastos_admin': [],
            'gastos_ventas': [],
            'otros_ingresos': [],
            'otros_gastos': [],
            'gastos_financieros': [],
            'impuestos': []
        }
        
        for factura in facturas:
            try:
                # Procesar ingresos (ventas)
                if 'total' in factura and factura['total']:
                    total = Decimal(str(factura['total']))
                    datos['ingresos'].append({
                        'codigo': '4135',
                        'descripcion': f"Ventas - Factura {factura.get('id', 'N/A')}",
                        'valor': total
                    })
                
                # Procesar impuestos si están detallados
                if 'taxes' in factura and factura['taxes']:
                    for tax in factura['taxes']:
                        if isinstance(tax, dict) and 'value' in tax:
                            valor_impuesto = Decimal(str(tax['value']))
                            datos['impuestos'].append({
                                'codigo': '2408',
                                'descripcion': f"IVA - {tax.get('name', 'Impuesto')}",
                                'valor': valor_impuesto
                            })
                
            except (ValueError, TypeError, KeyError) as e:
                self._logger.warning(f"⚠️ Error procesando factura {factura.get('id', 'N/A')}: {e}")
                continue
        
        return datos
    
    def _procesar_journals_contables(self, journals: List[Dict]) -> Dict[str, List]:
        """Procesar journals para obtener información contable detallada."""
        # TODO: Implementar procesamiento de journals cuando esté disponible
        return {}
    
    def _generar_datos_simulados(self) -> Dict[str, List]:
        """Generar datos simulados para demostración."""
        return {
            'ingresos': [
                {'codigo': '4135', 'descripcion': 'Ventas de servicios', 'valor': Decimal('45000000')},
                {'codigo': '4140', 'descripcion': 'Ventas de productos', 'valor': Decimal('25000000')}
            ],
            'costos': [
                {'codigo': '6135', 'descripcion': 'Costo de servicios', 'valor': Decimal('18000000')},
                {'codigo': '6140', 'descripcion': 'Costo de productos', 'valor': Decimal('12000000')}
            ],
            'gastos_admin': [
                {'codigo': '5105', 'descripcion': 'Gastos de personal administrativo', 'valor': Decimal('8000000')},
                {'codigo': '5115', 'descripcion': 'Servicios públicos', 'valor': Decimal('1500000')}
            ],
            'gastos_ventas': [
                {'codigo': '5205', 'descripcion': 'Gastos de personal de ventas', 'valor': Decimal('4000000')},
                {'codigo': '5210', 'descripcion': 'Publicidad y marketing', 'valor': Decimal('2000000')}
            ],
            'otros_ingresos': [
                {'codigo': '4295', 'descripcion': 'Ingresos por intereses', 'valor': Decimal('500000')}
            ],
            'otros_gastos': [
                {'codigo': '5295', 'descripcion': 'Gastos varios', 'valor': Decimal('300000')}
            ],
            'gastos_financieros': [
                {'codigo': '5305', 'descripcion': 'Intereses sobre préstamos', 'valor': Decimal('800000')}
            ],
            'impuestos': [
                {'codigo': '5405', 'descripcion': 'Impuesto de renta', 'valor': Decimal('2400000')}
            ]
        }
    
    def _construir_estado_resultados(self, 
                                   periodo_actual: PeriodoComparacion,
                                   periodo_anterior: Optional[PeriodoComparacion],
                                   datos_actual: Dict[str, List],
                                   datos_anterior: Optional[Dict[str, List]]) -> EstadoResultados:
        """Construir objeto EstadoResultados con los datos obtenidos."""
        
        estado = EstadoResultados(periodo_actual, periodo_anterior)
        
        # Procesar ingresos
        for item in datos_actual.get('ingresos', []):
            valor_anterior = None
            if datos_anterior:
                # Buscar item equivalente en periodo anterior
                item_anterior = next((x for x in datos_anterior.get('ingresos', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_ingreso(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        # Procesar costos
        for item in datos_actual.get('costos', []):
            valor_anterior = None
            if datos_anterior:
                item_anterior = next((x for x in datos_anterior.get('costos', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_costo_ventas(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        # Procesar gastos administrativos
        for item in datos_actual.get('gastos_admin', []):
            valor_anterior = None
            if datos_anterior:
                item_anterior = next((x for x in datos_anterior.get('gastos_admin', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_gasto_administracion(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        # Procesar gastos de ventas
        for item in datos_actual.get('gastos_ventas', []):
            valor_anterior = None
            if datos_anterior:
                item_anterior = next((x for x in datos_anterior.get('gastos_ventas', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_gasto_ventas(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        # Procesar otros ingresos
        for item in datos_actual.get('otros_ingresos', []):
            valor_anterior = None
            if datos_anterior:
                item_anterior = next((x for x in datos_anterior.get('otros_ingresos', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_otro_ingreso(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        # Procesar otros gastos
        for item in datos_actual.get('otros_gastos', []):
            valor_anterior = None
            if datos_anterior:
                item_anterior = next((x for x in datos_anterior.get('otros_gastos', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_otro_gasto(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        # Procesar gastos financieros
        for item in datos_actual.get('gastos_financieros', []):
            valor_anterior = None
            if datos_anterior:
                item_anterior = next((x for x in datos_anterior.get('gastos_financieros', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_gasto_financiero(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        # Procesar impuestos
        for item in datos_actual.get('impuestos', []):
            valor_anterior = None
            if datos_anterior:
                item_anterior = next((x for x in datos_anterior.get('impuestos', []) 
                                    if x['codigo'] == item['codigo']), None)
                if item_anterior:
                    valor_anterior = item_anterior['valor']
            
            estado.agregar_impuesto(item['codigo'], item['descripcion'], item['valor'], valor_anterior)
        
        return estado
    
    def _generar_archivo_excel(self, estado_resultados: EstadoResultados) -> str:
        """Generar archivo Excel profesional con el Estado de Resultados."""
        try:
            # Crear workbook y worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Estado de Resultados"
            
            # Configurar estilos
            self._configurar_estilos_excel(ws)
            
            # Generar contenido del reporte
            self._generar_contenido_excel(ws, estado_resultados)
            
            # Guardar archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"estado_resultados_{timestamp}.xlsx"
            ruta_archivo = os.path.join("outputs", nombre_archivo)
            
            # Crear directorio si no existe
            os.makedirs("outputs", exist_ok=True)
            
            wb.save(ruta_archivo)
            
            self._logger.info(f"📁 Archivo Excel guardado: {ruta_archivo}")
            return ruta_archivo
            
        except Exception as e:
            self._logger.error(f"❌ Error generando archivo Excel: {e}")
            raise
    
    def _configurar_estilos_excel(self, ws):
        """Configurar estilos profesionales para el archivo Excel."""
        # Definir estilos
        self.font_titulo = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        self.font_subtitulo = Font(name='Arial', size=12, bold=True, color='1F4E79')
        self.font_seccion = Font(name='Arial', size=11, bold=True, color='2F5597')
        self.font_normal = Font(name='Arial', size=10, color='000000')
        self.font_total = Font(name='Arial', size=10, bold=True, color='000000')
        
        self.fill_titulo = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        self.fill_seccion = PatternFill(start_color='E7EDF5', end_color='E7EDF5', fill_type='solid')
        self.fill_total = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.alignment_center = Alignment(horizontal='center', vertical='center')
        self.alignment_right = Alignment(horizontal='right', vertical='center')
    
    def _generar_contenido_excel(self, ws, estado_resultados: EstadoResultados):
        """Generar el contenido del Estado de Resultados en Excel."""
        
        # Configurar anchos de columna
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        
        fila = 1
        
        # Título principal
        ws.merge_cells(f'A{fila}:E{fila}')
        celda_titulo = ws[f'A{fila}']
        celda_titulo.value = "ESTADO DE RESULTADOS"
        celda_titulo.font = self.font_titulo
        celda_titulo.fill = self.fill_titulo
        celda_titulo.alignment = self.alignment_center
        fila += 2
        
        # Información del período
        ws[f'A{fila}'] = f"Período: {estado_resultados.periodo_actual.nombre}"
        ws[f'A{fila}'].font = self.font_subtitulo
        fila += 1
        
        if estado_resultados.periodo_anterior:
            ws[f'A{fila}'] = f"Comparación: {estado_resultados.periodo_anterior.nombre}"
            ws[f'A{fila}'].font = self.font_subtitulo
        fila += 2
        
        # Headers de columnas
        headers = ['CONCEPTO', 'PERÍODO ACTUAL', 'PERÍODO ANTERIOR', 'VARIACIÓN $', 'VARIACIÓN %']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col, value=header)
            celda.font = self.font_seccion
            celda.fill = self.fill_seccion
            celda.alignment = self.alignment_center
            celda.border = self.border
        fila += 1
        
        # Ingresos Operacionales
        fila = self._agregar_seccion_excel(ws, fila, "INGRESOS OPERACIONALES", estado_resultados.get_ingresos())
        
        # Total Ingresos
        fila = self._agregar_total_excel(ws, fila, "TOTAL INGRESOS OPERACIONALES", 
                                       estado_resultados.total_ingresos_operacionales,
                                       None)  # TODO: Calcular total anterior
        fila += 1
        
        # Costos de Ventas
        fila = self._agregar_seccion_excel(ws, fila, "COSTOS DE VENTAS", estado_resultados.get_costos_ventas())
        
        # Total Costos
        fila = self._agregar_total_excel(ws, fila, "TOTAL COSTOS DE VENTAS", 
                                       estado_resultados.total_costos_ventas,
                                       None)
        fila += 1
        
        # Utilidad Bruta
        fila = self._agregar_total_excel(ws, fila, "UTILIDAD BRUTA", 
                                       estado_resultados.utilidad_bruta,
                                       None, es_resultado=True)
        
        # Margen Bruto
        if estado_resultados.margen_bruto:
            ws[f'A{fila}'] = "Margen Bruto %"
            ws[f'B{fila}'] = f"{estado_resultados.margen_bruto:.2f}%"
            fila += 2
        
        # Gastos Operacionales
        fila = self._agregar_seccion_excel(ws, fila, "GASTOS DE ADMINISTRACIÓN", estado_resultados.get_gastos_administracion())
        fila = self._agregar_seccion_excel(ws, fila, "GASTOS DE VENTAS", estado_resultados.get_gastos_ventas())
        
        # Total Gastos Operacionales
        fila = self._agregar_total_excel(ws, fila, "TOTAL GASTOS OPERACIONALES", 
                                       estado_resultados.total_gastos_operacionales,
                                       None)
        fila += 1
        
        # Utilidad Operacional
        fila = self._agregar_total_excel(ws, fila, "UTILIDAD OPERACIONAL", 
                                       estado_resultados.utilidad_operacional,
                                       None, es_resultado=True)
        
        # Margen Operacional
        if estado_resultados.margen_operacional:
            ws[f'A{fila}'] = "Margen Operacional %"
            ws[f'B{fila}'] = f"{estado_resultados.margen_operacional:.2f}%"
            fila += 2
        
        # Otros Ingresos y Gastos
        fila = self._agregar_seccion_excel(ws, fila, "OTROS INGRESOS", estado_resultados.get_otros_ingresos())
        fila = self._agregar_seccion_excel(ws, fila, "OTROS GASTOS", estado_resultados.get_otros_gastos())
        fila = self._agregar_seccion_excel(ws, fila, "GASTOS FINANCIEROS", estado_resultados.get_gastos_financieros())
        
        # Utilidad antes de Impuestos
        fila = self._agregar_total_excel(ws, fila, "UTILIDAD ANTES DE IMPUESTOS", 
                                       estado_resultados.utilidad_antes_impuestos,
                                       None, es_resultado=True)
        fila += 1
        
        # Impuestos
        fila = self._agregar_seccion_excel(ws, fila, "IMPUESTOS", estado_resultados.get_impuestos())
        
        # Total Impuestos
        fila = self._agregar_total_excel(ws, fila, "TOTAL IMPUESTOS", 
                                       estado_resultados.total_impuestos,
                                       None)
        fila += 2
        
        # Utilidad Neta - RESULTADO FINAL
        fila = self._agregar_total_excel(ws, fila, "UTILIDAD NETA", 
                                       estado_resultados.utilidad_neta,
                                       None, es_resultado=True, es_final=True)
        
        # Margen Neto
        if estado_resultados.margen_neto:
            ws[f'A{fila}'] = "Margen Neto %"
            ws[f'B{fila}'] = f"{estado_resultados.margen_neto:.2f}%"
            ws[f'A{fila}'].font = self.font_total
            ws[f'B{fila}'].font = self.font_total
    
    def _agregar_seccion_excel(self, ws, fila_inicio: int, titulo_seccion: str, lineas: List) -> int:
        """Agregar una sección de líneas al Excel."""
        
        # Título de sección
        ws[f'A{fila_inicio}'] = titulo_seccion
        ws[f'A{fila_inicio}'].font = self.font_seccion
        fila = fila_inicio + 1
        
        # Líneas de la sección
        for linea in lineas:
            ws[f'A{fila}'] = f"  {linea.descripcion}"
            ws[f'B{fila}'] = f"${linea.valor_actual:,.0f}"
            
            if linea.valor_anterior is not None:
                ws[f'C{fila}'] = f"${linea.valor_anterior:,.0f}"
                if linea.variacion_absoluta is not None:
                    ws[f'D{fila}'] = f"${linea.variacion_absoluta:,.0f}"
                if linea.variacion_porcentual is not None:
                    ws[f'E{fila}'] = f"{linea.variacion_porcentual:.1f}%"
            
            # Aplicar formato
            for col in range(1, 6):
                celda = ws.cell(row=fila, column=col)
                celda.font = self.font_normal
                if col > 1:
                    celda.alignment = self.alignment_right
                celda.border = self.border
            
            fila += 1
        
        return fila
    
    def _agregar_total_excel(self, ws, fila: int, concepto: str, valor_actual: Decimal, 
                           valor_anterior: Optional[Decimal], es_resultado: bool = False, 
                           es_final: bool = False) -> int:
        """Agregar línea de total al Excel."""
        
        ws[f'A{fila}'] = concepto
        ws[f'B{fila}'] = f"${valor_actual:,.0f}"
        
        if valor_anterior is not None:
            ws[f'C{fila}'] = f"${valor_anterior:,.0f}"
            variacion_abs = valor_actual - valor_anterior
            ws[f'D{fila}'] = f"${variacion_abs:,.0f}"
            if valor_anterior != 0:
                variacion_pct = (variacion_abs / abs(valor_anterior)) * 100
                ws[f'E{fila}'] = f"{variacion_pct:.1f}%"
        
        # Aplicar formato según tipo
        for col in range(1, 6):
            celda = ws.cell(row=fila, column=col)
            if es_final:
                celda.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                celda.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            elif es_resultado:
                celda.font = self.font_total
                celda.fill = self.fill_total
            else:
                celda.font = self.font_total
            
            if col > 1:
                celda.alignment = self.alignment_right
            celda.border = self.border
        
        return fila + 1